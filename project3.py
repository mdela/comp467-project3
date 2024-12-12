# oh my god, the amount of libraries i had to install to get set up on this thing.
import argparse
import re
import subprocess
from pymongo import MongoClient
import pandas as pd
import os
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage

# connect to mongodb to store baselight and xytech data
client = MongoClient('mongodb://localhost:27017/')
db = client['video_processing']
baselight_collection = db['baselight']
xytech_collection = db['xytech']

# i have folders for the thumbnails and the snippets
# the thumbnails folder i use to load pictures and then insert into the xls file.
# the snippets i use to just upload to frame.io when i'm finished.
os.makedirs("thumbnails", exist_ok=True)
os.makedirs("snippets", exist_ok=True)

# get the total duration of the video using ffprobe
def get_video_length(video_path):
    # use ffprobe to get the video duration in seconds
    cmd = [
        'ffprobe', '-i', video_path, '-show_entries', 'format=duration', '-v', 'quiet', '-of', 'csv=p=0'
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    duration_seconds = float(result.stdout.strip())  # extract duration from ffprobe output
    return duration_seconds

# convert frame number to timecode in hh:mm:ss:ff format
def frame_to_timecode(frame, fps):
    # calculate hours, minutes, seconds, and frames based on the frame number and fps
    hours = frame // (fps * 3600)
    minutes = (frame // (fps * 60)) % 60
    seconds = (frame // fps) % 60
    frames = frame % fps
    # return the formatted timecode as a string (e.g., 01:02:03:04)
    return f"{hours:02}:{minutes:02}:{seconds:02}:{frames:02}"

# process the xytech file to extract producer, operator, job info, and location mappings
def process_xytech_file(xytech_filename):
    location_map = {}
    with open(xytech_filename, 'r') as xytech_file:
        producer = ''
        operator = ''
        job = ''
        for line in xytech_file:
            line = line.strip()
            if line.startswith('Producer:'):
                producer = line.split(':', 1)[1].strip()
            elif line.startswith('Operator:'):
                operator = line.split(':', 1)[1].strip()
            elif line.startswith('Job:'):
                job = line.split(':', 1)[1].strip()
            elif line:
                # clean location paths to strip unnecessary prefixes
                stripped_location = re.sub(r'^/hpsans\d+/production', '', line)
                location_map[stripped_location] = line

    # store the extracted data into mongodb
    xytech_collection.insert_one({
        'producer': producer,
        'operator': operator,
        'job': job,
        'locations': [{'stripped': k, 'full': v} for k, v in location_map.items()]
    })
    # return location mappings for later use
    return location_map

# process the baselight file and map frames to their corresponding locations
def process_baselight_file(baselight_filename, location_map):
    frames_locations = {}

    with open(baselight_filename, 'r') as baselight_file:
        for line in baselight_file:
            line = line.strip()
            if not line:
                continue

            # split line into location and frames
            parts = line.split()
            location_part = parts[0]
            frame_numbers = parts[1:]

            # strip the location and check if it's in the location map
            stripped_location = re.sub(r'^/baselightfilesystem1', '', location_part)

            if stripped_location in location_map:
                location_fixed = location_map[stripped_location]

                # map each frame to its corresponding location
                for frame in frame_numbers:
                    if frame.isdigit():
                        frames_locations[int(frame)] = location_fixed

    # sort frames and merge consecutive ones that share the same location
    frames_sorted = sorted(frames_locations.items())
    i = 0
    while i < len(frames_sorted):
        start_frame = frames_sorted[i][0]
        location = frames_sorted[i][1]
        end_frame = start_frame

        # merge consecutive frames if they share the same location
        while i + 1 < len(frames_sorted) and frames_sorted[i + 1][0] == end_frame + 1 and frames_sorted[i + 1][1] == location:
            end_frame = frames_sorted[i + 1][0]
            i += 1

        # insert either a single frame or a frame range into the database
        if start_frame == end_frame:
            baselight_collection.insert_one({
                'location': location,
                'frame': str(start_frame)
            })
        else:
            baselight_collection.insert_one({
                'location': location,
                'frame': f"{start_frame}-{end_frame}"
            })

        i += 1

# create a thumbnail for a specific frame in the video
def generate_thumbnail(video_filename, frame, output_path):
    cmd = [
        'ffmpeg', '-i', video_filename, '-vf', f"select=gte(n\,{frame})", '-vframes', '1',
        '-s', '96x74', output_path, '-y'
    ]
    # use ffmpeg to generate the thumbnail image from the specified frame
    subprocess.run(cmd, capture_output=True)

# filter frames, generate thumbnails, write to an XLS file, and create video snippets
# okay, so i tried to make it simpler. both the unused and used frames are in a single xls file
# one sheet is the used frame ranges (frames to fix) and the other is the unused frames and ranges
def filter_and_write_xls_and_snippets(video_filename, xls_filename, fps=24):
    video_length_seconds = get_video_length(video_filename)
    total_frames = int(video_length_seconds * fps)

    # retrieve metadata from xytech collection (producer, operator, job)
    xytech_data = xytech_collection.find_one()
    producer = xytech_data.get('producer', '') if xytech_data else ''
    operator = xytech_data.get('operator', '') if xytech_data else ''
    job = xytech_data.get('job', '') if xytech_data else ''

    # create a new workbook and add sheets for frames to fix and frames not used
    wb = Workbook()
    ws_frames_to_fix = wb.active
    ws_frames_to_fix.title = "Frames to Fix"
    ws_not_used = wb.create_sheet("Frames Not Used")

    # write metadata, or pretty much the xytech stuff, to the "Frames to Fix" sheet
    ws_frames_to_fix.append(["Producer", producer])
    ws_frames_to_fix.append(["Operator", operator])
    ws_frames_to_fix.append(["Job", job])
    ws_frames_to_fix.append([])
    ws_frames_to_fix.append(["Location", "Frames to Fix", "Timecode", "Thumbnail"])

    # write headers to "Frames Not Used" sheet
    ws_not_used.append(["Location", "Frames to Fix", "Timecode"])

    # loop through baselight collection records in the database and process each frame
    entries = list(baselight_collection.find({}))
    for entry in entries:
        frame_data = entry['frame']
        location = entry['location']

        # check if it's a range of frames
        if '-' in frame_data:
            start, end = map(int, frame_data.split('-'))
            timecode = f"{frame_to_timecode(start, fps)} - {frame_to_timecode(end, fps)}"

            if start <= total_frames and end <= total_frames:  # valid range
                # add to "Frames to Fix" sheet
                ws_frames_to_fix.append([location, frame_data, timecode])
                middle_frame = (start + end) // 2
                thumbnail_path = f"thumbnails/thumb_{start}_{end}.jpg"
                generate_thumbnail(video_filename, middle_frame, thumbnail_path)
                img = OpenpyxlImage(thumbnail_path)
                ws_frames_to_fix.add_image(img, f"D{ws_frames_to_fix.max_row}")

                # create video snippet for the frame range
                snippet_path = f"snippets/{start}-{end}.mp4"
                cmd = [
                    'ffmpeg', '-i', video_filename,
                    '-ss', f"{start / fps}", '-to', f"{end / fps}",
                    '-c', 'copy', snippet_path
                ]
                subprocess.run(cmd, capture_output=True)
            else:  # out-of-range frame range
                ws_not_used.append([location, frame_data, timecode])
        else:
            # process individual frame
            frame = int(frame_data)
            timecode = frame_to_timecode(frame, fps)

            if frame <= total_frames:  # valid individual frame
                ws_not_used.append([location, frame_data, timecode])
            else:  # out-of-range individual frame
                ws_not_used.append([location, frame_data, timecode])

    # save the generated XLS file
    wb.save(xls_filename)
    print(f"XLS file created with valid and unused frame data: {xls_filename}")

# pretty much where the program handles all of the arguments
def main():
    parser = argparse.ArgumentParser(description="process baselight files, calculate timecodes, and generate xls/snippets.")
    parser.add_argument('--xytech', required=True, help="path to the xytech file")
    parser.add_argument('--baselight', required=True, help="path to the baselight file")
    parser.add_argument('--process', help="path to the video file for xls/snippet creation")
    parser.add_argument('--outputXLS', help="path to save the xls file")
    args = parser.parse_args()

    # process xytech file to get location map
    location_map = process_xytech_file(args.xytech)

    # process baselight file and store frame-to-location mappings in the database
    process_baselight_file(args.baselight, location_map)

    # generate xls and snippets if both video file and output path are provided
    if args.process and args.outputXLS:
        filter_and_write_xls_and_snippets(args.process, args.outputXLS)

if __name__ == '__main__':
    main()
